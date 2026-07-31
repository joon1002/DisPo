#!/usr/bin/env python3
"""
HotpotQA full-corpus RobustRAG evaluation for five N=4 attack files.

Pipeline per query:
  clean Contriever top-50 cache + N=4 poison docs -> Contriever top-5
  -> ND: Mistral-7B generation
  -> RR: RobustRAG KeywordAgg with Mistral-7B
  -> RD+RR: HotpotQA multihop RAGDefender -> RobustRAG KeywordAgg with Mistral-7B
"""

import argparse
import ast
import gc
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import torch
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from sentence_transformers import SentenceTransformer
from tqdm import tqdm
from transformers import AutoModel, AutoTokenizer

_EVAL_ROOT = Path(__file__).resolve().parent
_ROOT = _EVAL_ROOT.parent
_DATA_ROOT = os.environ.get("DIPOISON_DATA_ROOT", "/path/to")
if str(_EVAL_ROOT) not in sys.path:
    sys.path.insert(0, str(_EVAL_ROOT))

from hotpotqa_multihop_ragdef_v2_eval import (  # noqa: E402
    GENERATOR_MODELS,
    HFChatGenerator,
    check_acc,
    check_asr,
    ragdefender_multihop,
    wrap_prompt,
)
from main_dipoison_fullcorpus_ragdef import contriever_encode  # noqa: E402


ATTACKS = {
    "PoisonedRAG": str(_ROOT / "data/attackbaselines_pd/PoisonedRAG/hotpotqa/poisonedrag4_hotpot100.csv"),
    "Joint-GCG": str(_ROOT / "data/attackbaselines_pd/jointgcg/hotpotqa/hotpotqa_origin_jointgcg_v2_n4.csv"),
    "Confundo": str(_ROOT / "data/attackbaselines_pd/confundo/hotpotqa/confundo_hotpotqa_N4.csv"),
    "RAGParadox": str(_ROOT / "data/attackbaselines_pd/RAGParadox/hotpotqa/hotpotqa_ragparadox_n4.csv"),
    "DiPoison": str(_ROOT / "data/attackbaselines_pd/DiPoison/hotpotqa/dipoison4_hotpot100.csv"),
}

_ISOLATION_PROMPT = (
    "Context information is below.\n"
    "---------------------\n"
    "{context}\n"
    "---------------------\n"
    "Given the context information and not prior knowledge, "
    "answer the query with only keywords.\n"
    "If there is no relevant information, just say \"I don't know\".\n"
    "Query: {query}\n"
    "Answer:"
)

_HINT_PROMPT = (
    "Word suggestion is below.\n"
    "---------------------\n"
    "{hints}\n"
    "---------------------\n"
    "Given the word suggestion provided by experts, concisely answer the query.\n"
    "Query: {query}\n"
    "Answer:"
)

_STOP_WORDS = {
    "a", "an", "and", "are", "as", "at", "be", "been", "by", "for", "from",
    "has", "have", "he", "her", "his", "i", "in", "is", "it", "its", "of",
    "on", "or", "she", "that", "the", "their", "there", "they", "this", "to",
    "was", "were", "which", "who", "with", "yes", "no", "not", "known",
    "answer", "query", "context", "information",
}


def extract_keywords_light(text):
    tokens = re.findall(r"[A-Za-z0-9][A-Za-z0-9'\\-]*", str(text).lower())
    keywords = {
        t.strip("'-.")
        for t in tokens
        if len(t.strip("'-.") ) > 1 and t.strip("'-.") not in _STOP_WORDS
    }
    for match in re.findall(r"(?:[A-Z][A-Za-z0-9'\\-]+(?:\\s+|$)){2,}", str(text)):
        phrase = " ".join(match.lower().split())
        if phrase and phrase not in _STOP_WORDS:
            keywords.add(phrase)
    return keywords


def robustrag_keyword_agg(question, topk_docs, llm, alpha=0.3, beta=3, abstention_threshold=1):
    individual_responses = []
    for doc in topk_docs:
        prompt = _ISOLATION_PROMPT.format(context=doc, query=question)
        try:
            resp = llm.query(prompt)
        except Exception:
            resp = ""
        individual_responses.append(resp)

    valid_responses = [r for r in individual_responses if "i don't" not in r.lower()]
    if len(valid_responses) < abstention_threshold:
        return "I don't know.", individual_responses, ""

    token_counter = Counter()
    for resp in valid_responses:
        for phrase in extract_keywords_light(resp):
            token_counter[phrase] += 1

    count_threshold = min(beta, alpha * len(valid_responses))
    filtered = {
        t: c for t, c in token_counter.items()
        if c >= count_threshold and t not in _STOP_WORDS
    }
    sorted_tokens = sorted(filtered.items(), key=lambda x: (len(x[0]), x[0]), reverse=True)
    hints = ", ".join(t for t, _ in sorted_tokens)

    if not hints:
        fallback_prompt = f"Answer the query concisely.\nQuery: {question}\nAnswer:"
        try:
            final_response = llm.query(fallback_prompt)
        except Exception:
            final_response = ""
    else:
        hint_prompt = _HINT_PROMPT.format(hints=hints, query=question)
        try:
            final_response = llm.query(hint_prompt)
        except Exception:
            final_response = ""
    return final_response, individual_responses, hints


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--corpus_path", default=f"{_DATA_ROOT}/datasets/hotpotqa/corpus.jsonl")
    p.add_argument("--clean_topn_cache", default=str(_EVAL_ROOT / "clean_topn_cache/hotpotqa_5attacks_top50/contriever_top50.pt"))
    p.add_argument("--output_dir", default=str(_EVAL_ROOT / "results/hotpotqa_mistral_robustrag_multihop_ragdef"))
    p.add_argument("--substring_dir", default=str(_EVAL_ROOT / "results/hotpotqa_mistral_robustrag_multihop_ragdef/substring_result"))
    p.add_argument("--defense_model", default="paraphrase-MiniLM-L6-v2")
    p.add_argument("--generator_model", default=GENERATOR_MODELS["mistral"])
    p.add_argument("--gpu_id", type=int, default=0)
    p.add_argument("--adv_per_query", type=int, default=4)
    p.add_argument("--top_k", type=int, default=5)
    p.add_argument("--seed", type=int, default=12)
    p.add_argument("--max_new_tokens", type=int, default=150)
    p.add_argument("--rr_alpha", type=float, default=0.3)
    p.add_argument("--rr_beta", type=float, default=3.0)
    p.add_argument("--rr_abstention", type=int, default=1)
    p.add_argument("--local_files_only", action="store_true")
    return p.parse_args()


def clean_str(s):
    s = str(s).strip()
    if len(s) > 1 and s[-1] == ".":
        s = s[:-1]
    return s.lower()


def load_attack_csv(path, adv_per_query):
    df = pd.read_csv(path)
    rows = []
    if "adv_texts" in df.columns:
        for _, row in df.iterrows():
            docs = ast.literal_eval(row["adv_texts"])
            rows.append({
                "query": str(row["question"]).strip(),
                "target_answer": str(row["incorrect answer"]).strip(),
                "correct_answer": str(row["correct answer"]).strip(),
                "poison_docs": [str(d).strip() for d in docs[:adv_per_query]],
            })
        return rows

    doc_cols = [c for c in ["doc0_seed", "doc1", "doc2", "doc3", "doc4", "doc5", "doc6"] if c in df.columns]
    for _, row in df.iterrows():
        docs = [
            str(row[c]).strip()
            for c in doc_cols
            if pd.notna(row[c]) and str(row[c]).strip()
        ][:adv_per_query]
        if not docs:
            continue
        rows.append({
            "query": str(row["query"]).strip(),
            "target_answer": str(row["target_answer"]).strip(),
            "correct_answer": str(row["correct_answer"]).strip(),
            "poison_docs": docs,
        })
    return rows


def save_ragparadox_substring_excel(detail_df, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "substring_hotpotqa_ragparadox_n4_mistral7b_robustrag_rdrr_fullcorpus_contriever_top5.xlsx"

    excel_df = pd.DataFrame({
        "#": range(1, len(detail_df) + 1),
        "쿼리": detail_df["query"],
        "Target Answer": detail_df["target_answer"],
        "Correct Answer": detail_df["correct_answer"],
        "ND 최종 답변": detail_df["nd_response"],
        "ND 공격 성공/실패": detail_df["nd_asr"].astype(bool).astype(int),
        "RobustRAG 최종 답변": detail_df["rr_response"],
        "RobustRAG 공격 성공/실패": detail_df["rr_asr"].astype(bool).astype(int),
        "RD+RobustRAG 최종 답변": detail_df["rd_rr_response"],
        "RD+RobustRAG 공격 성공/실패": detail_df["rd_rr_asr"].astype(bool).astype(int),
    })

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        excel_df.to_excel(writer, sheet_name="Human Eval", index=False)

    wb = load_workbook(out_path)
    ws = wb["Human Eval"]
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 6, "B": 80, "C": 80, "D": 24,
        "E": 100, "F": 18, "G": 100, "H": 22,
        "I": 100, "J": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(out_path)
    return out_path


def main():
    args = parse_args()
    if "CUDA_VISIBLE_DEVICES" not in os.environ:
        os.environ["CUDA_VISIBLE_DEVICES"] = str(args.gpu_id)
    torch.manual_seed(args.seed)
    np.random.seed(args.seed)
    torch.cuda.set_device(0 if torch.cuda.device_count() == 1 else args.gpu_id)
    device = f"cuda:{torch.cuda.current_device()}"

    out_dir = Path(args.output_dir) / datetime.now().strftime("run_%Y_%m_%d_%H_%M_%S")
    out_dir.mkdir(parents=True, exist_ok=True)
    log_path = out_dir / "run.log"

    def log(msg):
        print(msg, flush=True)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(str(msg) + "\n")

    log(f"[config] device={device}")
    log(f"[config] top_k={args.top_k} adv_per_query={args.adv_per_query}")
    log(f"[config] retriever=facebook/contriever")
    log(f"[config] defense_model={args.defense_model}")
    log(f"[config] generator=mistral-7b ({args.generator_model})")
    log(f"[config] clean_topn_cache={args.clean_topn_cache}")
    log(f"[config] rr_alpha={args.rr_alpha} rr_beta={args.rr_beta} rr_abstention={args.rr_abstention}")

    log("[load] clean Contriever top-N cache")
    cache = torch.load(args.clean_topn_cache, map_location="cpu", weights_only=True)
    meta = cache.get("meta", {})
    if meta.get("dataset") != "hotpotqa" or meta.get("retrieval_model") != "contriever":
        raise ValueError(f"Unexpected cache meta: {meta}")
    if int(meta.get("top_n", 0)) < args.top_k:
        raise ValueError(f"cache top_n={meta.get('top_n')} < top_k={args.top_k}")
    query_to_row = {str(q).strip(): i for i, q in enumerate(cache["queries"])}
    top_indices = cache["top_indices"].long()
    top_scores = cache["top_scores"].float()
    log(f"[load] cache queries={len(query_to_row)} top_n={meta.get('top_n')}")

    log(f"[load] corpus text: {args.corpus_path}")
    corpus_texts = []
    with open(args.corpus_path, encoding="utf-8") as f:
        for line in f:
            corpus_texts.append(json.loads(line).get("text", ""))
    log(f"[load] corpus passages={len(corpus_texts):,}")

    log("[load] Contriever")
    ctv_tok = AutoTokenizer.from_pretrained("facebook/contriever", local_files_only=args.local_files_only)
    ctv_mod = AutoModel.from_pretrained(
        "facebook/contriever",
        torch_dtype=torch.float32,
        local_files_only=args.local_files_only,
    ).to(device)
    ctv_mod.eval()

    def encode_ctv(texts):
        return contriever_encode(texts, ctv_mod, ctv_tok, device, batch_size=64)

    log(f"[load] multihop RAGDefender defense model: {args.defense_model}")
    defense_model = SentenceTransformer(args.defense_model)

    log(f"[load] Mistral generator: {args.generator_model}")
    llm = HFChatGenerator(
        args.generator_model,
        device,
        max_new_tokens=args.max_new_tokens,
        local_files_only=args.local_files_only,
    )
    log(f"[load] ready. GPU memory={torch.cuda.memory_allocated()/1e9:.1f}GB")

    def retrieve_topk(query, poison_docs):
        row_idx = query_to_row.get(str(query).strip())
        if row_idx is None:
            raise KeyError(f"query not in cache: {query}")

        clean_idx = top_indices[row_idx]
        clean_scores = top_scores[row_idx]
        q_emb = encode_ctv([query]).to(device).half()
        p_emb = encode_ctv(poison_docs).to(device).half()
        adv_scores = torch.mm(p_emb, q_emb.T).squeeze(1).float().cpu()

        all_scores = torch.cat([clean_scores, adv_scores], dim=0)
        order = all_scores.topk(args.top_k).indices.tolist()
        docs, is_poison = [], []
        n_clean = clean_idx.numel()
        for idx in order:
            if idx < n_clean:
                docs.append(corpus_texts[int(clean_idx[idx])])
                is_poison.append(False)
            else:
                docs.append(poison_docs[idx - n_clean])
                is_poison.append(True)
        return docs, is_poison

    summary_rows = []
    all_detail_frames = []
    run_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for attack_name, csv_path in ATTACKS.items():
        log(f"\n[attack] {attack_name}: {csv_path}")
        rows = load_attack_csv(csv_path, args.adv_per_query)
        rows = [r for r in rows if r["query"] in query_to_row]
        log(f"[attack] valid queries={len(rows)}")

        nd_asr = nd_acc = 0
        rr_asr = rr_acc = 0
        rd_rr_asr = rd_rr_acc = 0
        poison_top5 = rd_docs_total = 0
        detail_rows = []

        for q_idx, row in enumerate(tqdm(rows, desc=attack_name, dynamic_ncols=True)):
            query = row["query"]
            target = row["target_answer"]
            correct = row["correct_answer"]
            poison_docs = row["poison_docs"][:args.adv_per_query]

            top_docs, top_is_poison = retrieve_topk(query, poison_docs)
            poison_top5 += sum(top_is_poison)
            clean_docs_for_prompt = [clean_str(d) for d in top_docs]

            nd_resp = llm.query(wrap_prompt(query, clean_docs_for_prompt))
            nd_is_asr = check_asr(target, nd_resp)
            nd_is_acc = check_acc(correct, nd_resp)
            nd_asr += int(nd_is_asr)
            nd_acc += int(nd_is_acc)

            rr_resp, _, rr_hints = robustrag_keyword_agg(
                question=query,
                topk_docs=clean_docs_for_prompt,
                llm=llm,
                alpha=args.rr_alpha,
                beta=args.rr_beta,
                abstention_threshold=args.rr_abstention,
            )
            rr_is_asr = check_asr(target, rr_resp)
            rr_is_acc = check_acc(correct, rr_resp)
            rr_asr += int(rr_is_asr)
            rr_acc += int(rr_is_acc)

            rd_docs = ragdefender_multihop(top_docs, defense_model)
            rd_docs_total += len(rd_docs)
            rd_rr_resp, _, rd_rr_hints = robustrag_keyword_agg(
                question=query,
                topk_docs=[clean_str(d) for d in rd_docs],
                llm=llm,
                alpha=args.rr_alpha,
                beta=args.rr_beta,
                abstention_threshold=args.rr_abstention,
            ) if rd_docs else ("", [], "")
            rd_rr_is_asr = check_asr(target, rd_rr_resp) if rd_rr_resp else False
            rd_rr_is_acc = check_acc(correct, rd_rr_resp) if rd_rr_resp else False
            rd_rr_asr += int(rd_rr_is_asr)
            rd_rr_acc += int(rd_rr_is_acc)

            detail_rows.append({
                "attack": attack_name,
                "query_index": q_idx,
                "query": query,
                "target_answer": target,
                "correct_answer": correct,
                "poison_in_top5": sum(top_is_poison),
                "rd_num_docs": len(rd_docs),
                "nd_response": nd_resp,
                "nd_asr": bool(nd_is_asr),
                "nd_acc": bool(nd_is_acc),
                "rr_response": rr_resp,
                "rr_asr": bool(rr_is_asr),
                "rr_acc": bool(rr_is_acc),
                "rr_hints": rr_hints,
                "rd_rr_response": rd_rr_resp,
                "rd_rr_asr": bool(rd_rr_is_asr),
                "rd_rr_acc": bool(rd_rr_is_acc),
                "rd_rr_hints": rd_rr_hints,
            })

            if (q_idx + 1) % 10 == 0:
                n_now = q_idx + 1
                log(
                    f"  [{n_now}/{len(rows)}] "
                    f"ND-ASR={nd_asr/n_now:.1%} RR-ASR={rr_asr/n_now:.1%} "
                    f"RD+RR-ASR={rd_rr_asr/n_now:.1%}"
                )
            gc.collect()
            torch.cuda.empty_cache()

        n = len(rows)
        summary = {
            "run_at": run_at,
            "attack": attack_name,
            "docs_csv": csv_path,
            "dataset": "hotpotqa",
            "retriever": "contriever",
            "top_k": args.top_k,
            "defense": "hotpotqa_multihop_ragdef_v2_eval.ragdefender_multihop",
            "defense_model": args.defense_model,
            "robustrag": "local RobustRAG KeywordAgg implementation",
            "robustrag_generator": "mistral-7b",
            "generator_model": args.generator_model,
            "num_queries": n,
            "nd_asr": round(nd_asr / n, 4),
            "nd_acc": round(nd_acc / n, 4),
            "rr_asr": round(rr_asr / n, 4),
            "rr_acc": round(rr_acc / n, 4),
            "rd_rr_asr": round(rd_rr_asr / n, 4),
            "rd_rr_acc": round(rd_rr_acc / n, 4),
            "rr_asr_drop": round((nd_asr - rr_asr) / n, 4),
            "rd_rr_asr_drop": round((nd_asr - rd_rr_asr) / n, 4),
            "avg_poison_top5": round(poison_top5 / n, 4),
            "avg_rd_docs": round(rd_docs_total / n, 4),
        }
        summary_rows.append(summary)

        detail_df = pd.DataFrame(detail_rows)
        detail_path = out_dir / f"details_{attack_name}.csv"
        detail_df.to_csv(detail_path, index=False)
        all_detail_frames.append(detail_df)

        if attack_name == "RAGParadox":
            xlsx_path = save_ragparadox_substring_excel(detail_df, args.substring_dir)
            log(f"[save] RAGParadox substring excel: {xlsx_path}")

        pd.DataFrame(summary_rows).to_csv(out_dir / "summary.csv", index=False)
        log(
            f"[result] {attack_name}: "
            f"ND-ASR={summary['nd_asr']*100:.1f}% "
            f"RR-ASR={summary['rr_asr']*100:.1f}% "
            f"RD+RR-ASR={summary['rd_rr_asr']*100:.1f}%"
        )
        log(f"[save] {detail_path}")
        gc.collect()
        torch.cuda.empty_cache()

    pd.DataFrame(summary_rows).to_csv(out_dir / "summary.csv", index=False)
    with open(out_dir / "summary.json", "w", encoding="utf-8") as f:
        json.dump(summary_rows, f, ensure_ascii=False, indent=2)
    if all_detail_frames:
        pd.concat(all_detail_frames, ignore_index=True).to_csv(out_dir / "details_all.csv", index=False)
    log(f"[save] {out_dir / 'summary.csv'}")
    log(f"[save] {out_dir / 'summary.json'}")
    log("[done]")


if __name__ == "__main__":
    main()
